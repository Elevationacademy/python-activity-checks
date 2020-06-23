import enum
import openpyxl as oxl

workbook_file = '/home/excel.xlsx'

wb_data = oxl.load_workbook(workbook_file, read_only=False, data_only=True)
wb_formulas = oxl.load_workbook(workbook_file, read_only=True, data_only=False)
ws = wb_formulas['EmployeesData']
ws_data = wb_data['EmployeesData']
row_count = ws.max_row


class EmployeesDataCols(enum.Enum):
    No = 0
    BirthDate = 1
    FirstName = 2
    LastName = 3
    Gender = 4
    HireDate = 5
    Title = 6
    Department = 7
    Salary = 8
    Seniority = 9


class EmployeeDetails:
    def __init__(self, row, raw_data=False):
        self.No = row[EmployeesDataCols.No.value].value
        self.BirthDate = row[EmployeesDataCols.BirthDate.value].value
        self.FirstName = str(row[EmployeesDataCols.FirstName.value].value)
        self.LastName = row[EmployeesDataCols.LastName.value].value
        self.Gender = row[EmployeesDataCols.Gender.value].value
        self.HireDate = row[EmployeesDataCols.HireDate.value].value
        if raw_data:
            return
        self.Title = row[EmployeesDataCols.Title.value].value
        self.Department = row[EmployeesDataCols.Department.value].value
        self.Salary = row[EmployeesDataCols.Salary.value].value

    #        self.Seniority = row[EmployeesDataCols.Title.value].value

    def __eq__(self, other):
        return self.No == other.No

    def __ne__(self, other):
        return not self == other

    def __le__(self, other):
        return self.No <= other.No

    def __gt__(self, other):
        return not self <= other

    def __ge__(self, other):
        return self.No >= other.No

    def __lt__(self, other):
        return not self >= other


class EmployeesDb:
    def __init__(self):
        self.collection = {}
        self.fetch_employees()

    def fetch_employees(self):
        ws_e = wb_data['Employees']
        for row in ws_e.iter_rows(min_row=2, max_row=ws_e.max_row):
            self.collection[row[0].value] = EmployeeDetails(row, True)

    def fetch_title(self):
        ws_e = wb_data['titles']
        for row in ws_e.iter_rows(min_row=2, max_row=ws_e.max_row):
            to_date = row[3].value
            if to_date.year == 9999: # only take the last title.
                self.collection[row[0].value].Title = row[1].value

    def fetch_salaries(self):
        ws_e = wb_data['salaries']
        for row in ws_e.iter_rows(min_row=2, max_row=ws_e.max_row):
            to_date = row[3].value
            if to_date.year == 9999: # only take the last salary
                self.collection[row[0].value].Salary = row[1].value


    def fetch_departments(self):
        ws_e = wb_data['departments']
        department_codes = {}
        for row in ws_e.iter_rows(min_row=2, max_row=ws_e.max_row):
            department_codes[row[0].value] = row[1].value

        ws_e = wb_data['dept_emp']
        for row in ws_e.iter_rows(min_row=2, max_row=ws_e.max_row):
            to_date = row[3].value
            if to_date.year == 9999: # only take the last salary
                self.collection[row[0].value].Department = department_codes[row[1].value]

employees_db = EmployeesDb()

