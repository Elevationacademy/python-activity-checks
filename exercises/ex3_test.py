import openpyxl as oxl
from .EmployeesWorksheetModel import *
import unittest


class TestEx3(unittest.TestCase):

    def __init__(self, *args, **kwargs):
        super(TestEx3, self).__init__(*args, **kwargs)
        self.emp_ws_data = wb_data['Employees']
        self.emp_ws_formula = wb_formulas['Employees']

    def checkHeader(self, name, col):
        cell = self.emp_ws_data[1][col]
        assert cell.value.strip() == name.strip(), f"expected column header on cell {cell.coordinate} should be {name} " \
                                                   f"but it is {cell.value}"

    def check_col_formula(self, col_no, formula_substr):
        for row in self.emp_ws_formula.iter_rows(min_row=2, max_row=row_count):
            cell = row[col_no]
            assert cell.data_type == 'f', f"cell {cell.coordinate} should be a formula"
            for f in formula_substr:
                assert f in cell.value, f"cell {cell.coordinate} formula should include {f} function"
             # todo: check that it is referencing the same row.
#            for r in ref_substr:
#                assert r in cell.value, f"cell {cell.coordinate} formula should reference to {r}"

    def test_FirstNameFirstLetter(self):
        self.checkHeader('FirstNameFirstLetter', 6)
        self.check_col_formula(6, ["LEFT"])
        for row in self.emp_ws_data.iter_rows(min_row=2, max_row=row_count):
            expected = row[2].value[0]
            cell = row[6]
            assert expected == cell.value, f"cell {cell.coordinate} expected value is {expected} but it " \
                                            f"is actual {cell.value}"

    def test_LastNameLastLetter(self):
        self.checkHeader('LastNameLastLetter', 7)
        self.check_col_formula(7, ["RIGHT"])
        for row in self.emp_ws_data.iter_rows(min_row=2, max_row=row_count):
            expected = row[3].value[-1]
            cell = row[7]
            assert expected == cell.value, f"cell {cell.coordinate} expected value is {expected} but it " \
                                            f"is actual {cell.value}"

    def test_TrimedEdgesName(self):
        self.checkHeader('TrimedEdgesName', 8)
        self.check_col_formula(8, ["MID"])
        for row in self.emp_ws_data.iter_rows(min_row=2, max_row=row_count):
            expected = row[2].value[1:-1] + " " + row[3].value[1:-1]
            cell = row[8]
            assert expected == cell.value, f"cell {cell.coordinate} expected value is {expected} but it " \
                                            f"is actual {cell.value}"

    def test_FirstLetterIndex(self):
        self.checkHeader('FirstLetterIndex', 9)
        self.check_col_formula(9, ["FIND"])
        for row in self.emp_ws_data.iter_rows(min_row=2, max_row=row_count):
            expected = row[3].value.lower().find(row[2].value[0].lower()) + 1
            cell = row[9]
            assert expected == cell.value, f"cell {cell.coordinate} expected value is {expected} but it " \
                                            f"is actual {cell.value}"


