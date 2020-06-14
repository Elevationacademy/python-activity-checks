import enum
import openpyxl as oxl

workbook_file = '../Solution/BikeStoreSample.xlsx'

wb_data = oxl.load_workbook(workbook_file, read_only=False, data_only=True)
wb_formulas = oxl.load_workbook(workbook_file, read_only=True, data_only=False)
ws = wb_formulas['OrderDetailsData']
ws_data = wb_data['OrderDetailsData']
row_count = ws.max_row


class BikeStoreSheetCols(enum.Enum):
    OrderId = 0
    ItemId = 1
    ProductName = 2
    ModelYear = 3
    ListPrice = 5
    Quantity = 6
    LineTotal = 7
    Discount = 8
    DiscountPerUnit = 9
    LineTotalAfterDiscount = 10
    OrderDate = 11
    SalesMan = 12
    OldModel = 15
