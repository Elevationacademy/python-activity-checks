import enum
import openpyxl as oxl

workbook_file = '/home/excel.xlsx'

wb_data = oxl.load_workbook(workbook_file, read_only=False, data_only=True)
wb_formulas = oxl.load_workbook(workbook_file, read_only=True, data_only=False)
ws = wb_formulas['vgsales']
ws_data = wb_data['vgsales']
row_count = ws.max_row

class VideoGameSalesSheetCols(enum.Enum):
    Rank = 0
    Name = 1
    Platform = 2
    Year = 3
    Genre = 4
    Publisher = 5
    NaSales = 6
    EuSales = 7
    JpSales = 8
    OtherSales = 9
    GlobalSales = 10


class VideoGameDetails:
    def __init__(self, row):
        self.Year = row[VideoGameSalesSheetCols.Year.value].value
        self.Genre = row[VideoGameSalesSheetCols.Genre.value].value
        self.Platform = str(row[VideoGameSalesSheetCols.Platform.value].value)
        self.Publisher = row[VideoGameSalesSheetCols.Publisher.value].value
        self.Rank = row[VideoGameSalesSheetCols.Rank.value].value

    def __eq__(self, other):
        (self.Year, self.Genre, self.Platform, self.Rank, self.Publisher) == (other.Year, other.Genre, other.Platform, other.Rank, other.Publisher)

    def __ne__(self, other):
        return not self == other

    def __le__(self, other):
        return (self.Year, self.Genre, self.Platform, self.Rank) <= (other.Year, other.Genre, other.Platform, other.Rank)

    def __gt__(self, other):
        return not self <= other

    def __ge__(self, other):
        return (self.Year, self.Genre, self.Platform, self.Rank) >= (other.Year, other.Genre, other.Platform, other.Rank)

    def __lt__(self, other):
        return not self >= other


