import openpyxl as oxl
from .videoGameWorksheetModel import *
import unittest


class TestEx4(unittest.TestCase):
    def __init__(self, *args, **kwargs):
        super(TestEx4, self).__init__(*args, **kwargs)
        self.worksheet_data = wb_data['PlatformByYear']
        self.worksheet_formula = wb_formulas['PlatformByYear']
        self.max_row = self.worksheet_data.max_row

        # create the raw data.
        self.expected_platform_sums = {}
        for row in ws_data.iter_rows(min_row=2, max_row=row_count):
            platform = str(row[VideoGameSalesSheetCols.Platform.value].value)
            year = row[VideoGameSalesSheetCols.Year.value].value
            if year not in self.expected_platform_sums.keys():
                self.expected_platform_sums[year] = {}
            if platform not in self.expected_platform_sums[year].keys():
                self.expected_platform_sums[year][platform] = row[VideoGameSalesSheetCols.GlobalSales.value].value
            else:
                self.expected_platform_sums[year][platform] += row[VideoGameSalesSheetCols.GlobalSales.value].value

    def test_PlatformColumn(self):
        formula_cell = self.worksheet_formula['A2']
        assert 'UNIQUE' in formula_cell.value, f"formula in cell {formula_cell.coordinate} should include UNIQUE in" \
                                               f" it's formula"
        assert 'vgsales!' in formula_cell.value, f"formula in cell {formula_cell.coordinate} should include reference to " \
                                                 f"vgsales tab"

    def test_YearsHeader(self):
        formula_cell = self.worksheet_formula['B1']
        assert 'UNIQUE' in formula_cell.value, f"formula in cell {formula_cell.coordinate} should include UNIQUE in" \
                                               f" it's formula"
        assert 'TRANSPOSE' in formula_cell.value, f"formula in cell {formula_cell.coordinate} should include TRANSPOSE in" \
                                                  f" it's formula"
        assert 'vgsales!' in formula_cell.value, f"formula in cell {formula_cell.coordinate} should include reference to " \
                                                 f"vgsales tab"

    def test_TotalIncomes(self):
        for row in self.worksheet_data.iter_rows(min_row=2, max_row=row_count):
            if row[0].value is None:
                break
            for cell in row[1:]:
                year = self.worksheet_data[1][cell.column - 1].value
                platform = str(self.worksheet_data[cell.row][0].value)
                #                print(cell.coordinate, year, platform)
                expected_val = 0 if platform not in self.expected_platform_sums[year] else \
                    self.expected_platform_sums[year][platform]
                assert expected_val == cell.value, f"income on Platform {platform} on year {year} should be {expected_val}" \
                                                   f"but it is {cell.value}"

    def test_TestLineChart(self):
        assert len(self.worksheet_data._charts) > 0, f"'PlatformByYear' does not contain any chart"
        assert len(self.worksheet_data._charts) < 2, f"'PlatformByYear' should have only 1 chart"
        assert self.worksheet_data._charts[0].tagname == 'scatterChart', f"'PlatformByYear' chart type should be 'line/Scatter chart'"

    def test_ConditionalFormatting(self):
        assert len(self.worksheet_data.conditional_formatting) > 0, f"There is no conditional formatting in tab ,PlatformByYear"