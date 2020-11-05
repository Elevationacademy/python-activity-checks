import openpyxl as oxl
from .videoGameWorksheetModel import *
import unittest

class TestEx5(unittest.TestCase):
    def __init__(self, *args, **kwargs):
        super(TestEx5, self).__init__(*args, **kwargs)
        self.worksheet_data = wb_data['GenrePerDecade']
        self.worksheet_formula = wb_formulas['GenrePerDecade']
        self.max_row = self.worksheet_data.max_row

        # create the raw data.
        self.expected_genres_avgs = {}
        expected_genres_counts = {}
        for row in ws_data.iter_rows(min_row=2, max_row=row_count):
            genre = row[VideoGameSalesSheetCols.Genre.value].value
            year = row[VideoGameSalesSheetCols.Year.value].value
            decade = year - year % 10
            if decade not in self.expected_genres_avgs.keys():
                self.expected_genres_avgs[decade] = {}
                expected_genres_counts[decade] = {}
            if genre not in self.expected_genres_avgs[decade].keys():
                self.expected_genres_avgs[decade][genre] = row[VideoGameSalesSheetCols.GlobalSales.value].value
                expected_genres_counts[decade][genre] = 1
            else:
                self.expected_genres_avgs[decade][genre] += row[VideoGameSalesSheetCols.GlobalSales.value].value
                expected_genres_counts[decade][genre] += 1

        for d in self.expected_genres_avgs:
            for g in self.expected_genres_avgs[d]:
                self.expected_genres_avgs[d][g] /= float(expected_genres_counts[d][g])


    def test_DecadeHeader(self):
        formula_cell = self.worksheet_formula['A2']
        assert 'UNIQUE' in formula_cell.value, f"formula in cell {formula_cell.coordinate} should include UNIQUE in" \
                                               f" it's formula"
        assert 'vgsales!' in formula_cell.value, f"formula in cell {formula_cell.coordinate} should include reference to " \
                                                 f"vgsales tab"

    def test_GenreCol(self):
        formula_cell = self.worksheet_formula['B1']
        assert 'UNIQUE' in formula_cell.value, f"formula in cell {formula_cell.coordinate} should include UNIQUE in" \
                                               f" it's formula"
        assert 'TRANSPOSE' in formula_cell.value, f"formula in cell {formula_cell.coordinate} should include TRANSPOSE in" \
                                                  f" it's formula"
        assert 'vgsales!' in formula_cell.value, f"formula in cell {formula_cell.coordinate} should include reference to " \
                                                 f"vgsales tab"

    def test_AverageIncomes(self):
        for row in self.worksheet_data.iter_rows(min_row=2, max_row=row_count):
            if row[0].value is None:
                break
            for cell in row[1:]:
                decade = self.worksheet_data[1][cell.column - 1].value
                genre = self.worksheet_data[cell.row][0].value
                expected_val = 0 if genre not in self.expected_genres_avgs[decade] \
                    else self.expected_genres_avgs[decade][genre]
                self.assertAlmostEqual(expected_val, cell.value, places=3,
                    msg=f"Average income in the {decade}'s for genre {genre} should be {expected_val} but it is {cell.value}")

    def test_BarCharts(self):
        assert len(self.worksheet_data._charts) > 0, f"'GenrePerDecade' does not contain any chart"
        assert len(self.worksheet_data._charts) < 2, f"'GenrePerDecade' should have only 1 chart"
        assert self.worksheet_data._charts[0].tagname == 'barChart', f"'GenrePerDecade' chart type should be 'bar charts'"
